from PyQt5.QtWidgets import *
from PyQt5.uic import loadUiType
from PyQt5.QtCore import Qt
import os
from os import path
import sys
import openpyxl
import xlrd
import pandas as pd
import numpy as np 

# Load the UI file
Ui_MainWindow, QMainWindow = loadUiType("interface.ui")

class MyMainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MyMainWindow, self).__init__()
        self.setupUi(self)

        self.setupUi(self)
        self.initUI()

    def initUI(self):
       
        self.P0 = []
        self.Q0 = []
        # Configure the buttons in the header_frame
        
        self.pushButton_2.clicked.connect(self.back_button_clicked)
        self.pushButton_3.clicked.connect(self.next_button_clicked)
        self.pushButton_9.clicked.connect(self.close_window)
        self.pushButton_8.clicked.connect(self.back_button_clicked)
        self.pushButton_10.clicked.connect(self.next_button_clicked)
        self.pushButton_11.clicked.connect(self.next_button_clicked)
        self.pushButton_4.clicked.connect(self.minimize_window)
        self.pushButton_6.clicked.connect(self.help_program)
        self.pushButton_5.clicked.connect(self.close_window)
        
        self.BT_import.clicked.connect(self.insert_excel_data)
        
        self.BT_Edit.clicked.connect(self.edit_selected_row)
        self.BT_Add.clicked.connect(self.add_row)
        self.BT_Delete.clicked.connect(self.delete_selected_row)
        self.BT_Export.clicked.connect(self.export_to_excel)
        self.BT_Edit_scenario.clicked.connect(self.edit_selected_SCENSRIO)
        self.BT_Add_scenario.clicked.connect(self.add_SCENSRIO)
        self.BT_Delete_scenario.clicked.connect(self.delete_selected_SCENSRIO)
        self.BT_load_scenario.clicked.connect(self.load_scenarios)
        self.BT_auto_scenario.clicked.connect(self.Auto_load_scenarios)
        self.pushButton_13.clicked.connect(self.toggle_size)
        #self.pushButton_recherche.clicked.connect(self.search_button_clicked)
        self.pushButton_recherche.clicked.connect(self.search_and_highlight)
        # Set window style
        self.setWindowFlags(Qt.Window | Qt.FramelessWindowHint)
        
        # Initialize current_page_index to 0
        self.current_page_index = 0

        # Set minimum and maximum window size
        self.setMinimumSize(1200, 800)  # Set the minimum size (width, height)
        self.tableWidget.itemChanged.connect(self.on_item_changed)
        
        # Initialize current_page_index to 0
        self.current_page_index = 0
        
        
        
        
############################################   scenarios   ############################################
    
    
    
    def toggle_size(self):
        if self.isMaximized():
            self.showNormal()
            
        else:
            self.showMaximized()
            
    def Auto_load_scenarios(self):
        def generate_scenarios(n):
            scenarios = []
            for i in range(2**n):
                binary_str = format(i, f'0{n}b')
                scenario = [int(bit) for bit in binary_str]
                scenarios.append(scenario)
            return scenarios

        n_interrupteurs = 16
        matrice_scenarios = generate_scenarios(n_interrupteurs)        

        for scenario in matrice_scenarios:
            row_position = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(row_position)
            
            for column, value in enumerate(scenario):
                item = QTableWidgetItem("close" if value == 1 else "open")
                self.tableWidget_2.setItem(row_position, column, item)

        QMessageBox.information(self, "Success", "Data inserted successfully.")     

    def load_scenarios(self):
        # Load scenarios from your data source
        matrice_scenarios=np.array([[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    [1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],[1,1,1,1,1,1,1,1,1,1,1,0,0,0,0,0],
                                    ])
        for scenario in matrice_scenarios:
            row_position = self.tableWidget_2.rowCount()
            self.tableWidget_2.insertRow(row_position)
            
            for column, value in enumerate(scenario):
                item = QTableWidgetItem("close" if value == 1 else "open")
                self.tableWidget_2.setItem(row_position, column, item)

        QMessageBox.information(self, "Success", "Data inserted successfully.") 
    

    def delete_scenario(self):
        selected_rows = sorted(set(index.row() for index in self.tableWidget_2.selectedIndexes()), reverse=True)
        for row in selected_rows:
            self.tableWidget_2.removeRow(row)

    def edit_scenario(self):
        selected_indexes = self.tableWidget_2.selectedIndexes()
        if len(selected_indexes) > 0:
            selected_row = selected_indexes[0].row()
            for column in range(self.tableWidget_2.columnCount()):
                item = QTableWidgetItem("Edited Value")
                self.tableWidget_2.setItem(selected_row, column, item)



            
    def edit_selected_SCENSRIO(self):
        selected_row = self.tableWidget_2.currentRow()
        if selected_row >= 0:
            for col in range(self.tableWidget_2.columnCount()):
                item = self.tableWidget_2.item(selected_row, col)
                if item is not None:
                    value, ok = QInputDialog.getText(self, "Edit Value", f"Edit value for column {col+1}", text=item.text())
                    if ok:
                        item.setText(value)

                        if col == 5:
                            self.P0[selected_row] = value
                        elif col == 6:
                            self.Q0[selected_row] = value
            self.update_doubleSpinBox_values()
            
    def add_SCENSRIO(self):
        row_count = self.tableWidget_2.rowCount()
        self.tableWidget_2.setRowCount(row_count + 1)

        self.P0.append("")
        self.Q0.append("")
        self.update_doubleSpinBox_values()
        
    def delete_selected_SCENSRIO(self):
        selected_row = self.tableWidget_2.currentRow()
        if selected_row >= 0:
            self.tableWidget_2.removeRow(selected_row)

            del self.P0[selected_row]
            del self.Q0[selected_row]
            self.update_doubleSpinBox_values()
    def search_button_clicked(self):
        search_value = self.lineEdit_search.text()  # Remplacez "lineEdit_search" par le nom de votre QLineEdit de recherche
        column_index = 0  # Mettez l'indice de la colonne que vous souhaitez rechercher ici
        self.search_and_visualize_row(search_value, column_index)
    def search_and_visualize_row(self, search_value, column_index):
        found_row = -1
        for row in range(self.tableWidget_2.rowCount()):
            item = self.tableWidget_2.item(row, column_index)
            if item is not None and item.text() == search_value:
                found_row = row
                break

        if found_row != -1:
            self.tableWidget_2.selectRow(found_row)
            self.tableWidget_2.scrollToItem(self.tableWidget_2.item(found_row, column_index))
            QMessageBox.information(self, "Search Result", f"Row with value '{search_value}' found at index {found_row}.")
        else:
            QMessageBox.warning(self, "Search Result", f"Value '{search_value}' not found in the specified column.")

    def highlight_row(self, row_number):
        if 0 <= row_number < self.tableWidget_2.rowCount():
            self.tableWidget_2.selectRow(row_number)
            self.tableWidget_2.setFocus()
        else:
            QMessageBox.warning(self, "Warning", f"Row {row_number} does not exist.")



    def search_and_highlight(self):
        row_number_text, ok = QInputDialog.getText(self, "Search Row", "Enter row number to search:")
        if ok:
            try:
                row_number = int(row_number_text)
                self.highlight_row(row_number)
            except ValueError:
                QMessageBox.warning(self, "Error", "Invalid row number. Please enter a valid integer.")


###############################################################################################################################

#####################################################    Page 2 : tableau1: Reseau IEEE 33 BUS    ##########################################################################
    def edit_selected_row(self):
        row_head = ['Branch Number',	'Sending Bus',	'Receiving Bus',	'Resistance',  'Reactance', 'P (kW)',	'Q (kVAr)']
        selected_row = self.tableWidget.currentRow()
        if selected_row >= 0:
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(selected_row, col)
                if item is not None:
                    value, ok = QInputDialog.getText(self, "Edit Value", f"Edit {row_head[col]}", text=item.text())
                    if ok:
                        item.setText(value)
                        if col == 5:
                            self.P0[selected_row] = value
                        elif col == 6:
                            self.Q0[selected_row] = value
            self.update_doubleSpinBox_values()
            
    def add_row(self):
        row_head = ['Branch Number',	'Sending Bus',	'Receiving Bus',	'Resistance',  'Reactance', 'P (kW)',	'Q (kVAr)']
        row_count = self.tableWidget.rowCount()
        self.tableWidget.setRowCount(row_count + 1)
        for col in range(self.tableWidget.columnCount()):
            item = self.tableWidget.item(row_count, col)
            if item is None:
                value, ok = QInputDialog.getText(self, "Add Value", f"add  {row_head[col]}")
                if ok:
                    item = QTableWidgetItem(value)
                    self.tableWidget.setItem(row_count, col, item)
                    if col == 5:
                        self.P0.append(value)
                    elif col == 6:
                        self.Q0.append(value)
        self.update_doubleSpinBox_values()
        
    def export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            data = []
            for row in range(self.tableWidget.rowCount()):
                row_data = []
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                data.append(row_data)
            df = pd.DataFrame(data, columns=['Branch Number',	'Sending Bus',	'Receiving Bus',	'Resistance',  'Reactance', 'P (kW)',	'Q (kVAr)'])
            df.to_excel(file_path, index=False)
            print("Excel file saved.")
        
    def delete_selected_row(self):
        selected_row = self.tableWidget.currentRow()
        if selected_row >= 0:
            self.tableWidget.removeRow(selected_row)
            del self.P0[selected_row]
            del self.Q0[selected_row]
            self.update_doubleSpinBox_values()
            
    def update_doubleSpinBox_values(self):
        for i, (p_value, q_value) in enumerate(zip(self.P0, self.Q0)):
            p_spinbox = getattr(self, f"P_{i+2}")
            q_spinbox = getattr(self, f"Q_{i+2}")
            p_spinbox.setValue(float(p_value) if p_value else 0.0)
            q_spinbox.setValue(float(q_value) if q_value else 0.0)

    def on_item_changed(self, item):
        row_index = item.row()
        col_index = item.column()
        value = item.text()
        if col_index == 0:
            if row_index < len(self.P0):
                self.P0[row_index] = value
        elif col_index == 1:
            if row_index < len(self.Q0):
                self.Q0[row_index] = value
        self.update_doubleSpinBox_values()          
############################### footer   & header  ###################################################################

    def back_button_clicked(self):
        if self.current_page_index > 0:
            self.current_page_index -= 1
            self.stackedWidget.setCurrentIndex(self.current_page_index)

    def next_button_clicked(self):
        total_pages = self.stackedWidget.count()
        if self.current_page_index < total_pages - 1:
            self.current_page_index += 1
            self.stackedWidget.setCurrentIndex(self.current_page_index)


            

    def insert_excel_data(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)", options=options)
        
        if not path.isfile(file_name):
            QMessageBox.warning(self, "Error", "Invalid file path.")
            return

        try:
            workbook = openpyxl.load_workbook(file_name)
            worksheet = workbook.active

            self.tableWidget.setRowCount(worksheet.max_row)
            self.tableWidget.setColumnCount(worksheet.max_column)

            self.P0.clear()
            self.Q0.clear()

            for row in worksheet.iter_rows():
                for cell in row:
                    col_index = cell.column - 1
                    row_index = cell.row - 1
                    item = QTableWidgetItem(str(cell.value))
                    self.tableWidget.setItem(row_index, col_index, item)

                    if col_index == 5:
                        self.P0.append(str(cell.value))
                    elif col_index == 6:
                        self.Q0.append(str(cell.value))

            QMessageBox.information(self, "Success", "Data inserted from Excel successfully.")
            self.update_doubleSpinBox_values()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred: {str(e)}")


    def close_window(self):
        self.close()

    def set_min_max_size(self):
        print("Button clicked: set_min_max_size")
        self.setMaximumSize(200, 200)

    def back_button_clicked(self):
        if self.current_page_index > 0:
            self.current_page_index -= 1
            self.stackedWidget.setCurrentIndex(self.current_page_index)

    def minimize_window(self):
        self.setWindowState(Qt.WindowMinimized)

    def help_program(self):
        help_message = "This is a sample help message.\nYou can customize it with your own help content."
        QMessageBox.information(self, "Help", help_message, QMessageBox.Ok)


        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyMainWindow()
    window.show()
    sys.exit(app.exec_())
